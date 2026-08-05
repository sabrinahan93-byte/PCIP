from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from pcip.utils.time import now_sgt_str
import os


class DashboardWorkbook:

    def __init__(self):

        self.file_path = (
            "output/Job_Dashboard.xlsx"
        )


    def initialize(self):

        os.makedirs(
            "output",
            exist_ok=True
        )


        if os.path.exists(
            self.file_path
        ):

            return


        wb = Workbook()


        ws = wb.active
        ws.title = "Jobs"


        ws.append([

            "Job ID",
            "Company",
            "Job Title",
            "Location",
            "Work Mode",
            "Match Score",
            "Job URL",
            "Last Seen",
            "Last Notified",
            "Pipeline Status",
            "Applied Date",
            "Notes"

        ])


        ws2 = wb.create_sheet(
            "Companies"
        )


        ws2.append([

            "Company",
            "Enabled",
            "Tier",
            "Career URL",
            "Last Scan",
            "Scan Status",
            "Notes"

        ])


        ws3 = wb.create_sheet(
            "Run_Log"
        )


        ws3.append([

            "Run Time",
            "Companies Scanned",
            "New Jobs",
            "Updated Jobs",
            "Closed Jobs",
            "Failed Companies",
            "Duration",
            "Result"

        ])


        ws4 = wb.create_sheet(
            "Run_Detail"
        )


        ws4.append([

            "Run Time",
            "Change Type",
            "Company",
            "Job URL",
            "Error"

        ])



        for sheet in wb:

            for cell in sheet[1]:

                cell.font = Font(
                    bold=True
                )


        wb.save(
            self.file_path
        )



    def sync_companies(
        self,
        companies
    ):

        wb = load_workbook(
            self.file_path
        )


        ws = wb["Companies"]


        existing = set()


        for row in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            if row[0]:

                existing.add(
                    row[0]
                )


        for company in companies:


            company_name = company.get(
                "Company",
                ""
            )


            if company_name not in existing:


                ws.append([

                    company_name,

                    company.get(
                        "Enabled",
                        True
                    ),

                    company.get(
                        "Tier",
                        ""
                    ),

                    company.get(
                        "Career URL",
                        ""
                    ),

                    "",

                    "",

                    ""

                ])


        wb.save(
            self.file_path
        )



    def write_jobs(
        self,
        jobs
    ):

        wb = load_workbook(
            self.file_path
        )


        ws = wb["Jobs"]

        detail_ws = wb["Run_Detail"]


        now = now_sgt_str()


        existing_jobs = {}


        for row in ws.iter_rows(
            min_row=2
        ):

            url = row[6].value


            if url:

                existing_jobs[url] = {

                    "row": row[0].row,

                    "pipeline_status": row[9].value,

                    "applied_date": row[10].value,

                    "notes": row[11].value

                }



        new_count = 0

        updated_count = 0



        for job in jobs:


            url = job.get(
                "Job URL",
                ""
            )


            company = job.get(
                "Company",
                ""
            )


            title = job.get(
                "Job Title",
                ""
            )


            score = job.get(
                "Match Score",
                0
            )



            if url in existing_jobs:


                row_num = existing_jobs[url]["row"]


                ws.cell(
                    row_num,
                    2
                ).value = company


                ws.cell(
                    row_num,
                    3
                ).value = title


                ws.cell(
                    row_num,
                    6
                ).value = score


                ws.cell(
                    row_num,
                    8
                ).value = now

                
                # Preserve manual fields
                ws.cell(
                    row_num,
                    10
                ).value = existing_jobs[url]["pipeline_status"]
                

                ws.cell(
                    row_num,
                    11
                ).value = existing_jobs[url]["applied_date"]
                

                ws.cell(
                    row_num,
                    12
                ).value = existing_jobs[url]["notes"]

                

                change_type = "UPDATED"


                updated_count += 1



            else:


                new_count += 1


                date_id = now.replace(
                    "-",
                    ""
                ).replace(
                    ":",
                    ""
                ).replace(
                    " ",
                    ""
                )


                job_id = (

                    "JOB-"

                    +

                    date_id[:8]

                    +

                    "-"

                    +

                    str(
                        ws.max_row
                    ).zfill(4)

                )


                ws.append([

                    job_id,

                    company,

                    title,

                    job.get(
                        "Location",
                        ""
                    ),

                    job.get(
                        "Work Mode",
                        ""
                    ),

                    score,

                    url,

                    now,

                    "",

                    "New",

                    "",

                    ""

                ])


                change_type = "NEW"



            detail_ws.append([

                now,

                change_type,

                company,

                url,

                ""

            ])



        wb.save(
            self.file_path
        )


        return {

            "new":
                new_count,

            "updated":
                updated_count

        }



    def write_scan_failure(
        self,
        failures
    ):

        wb = load_workbook(
            self.file_path
        )


        ws = wb["Run_Detail"]


        now = now_sgt_str()


        for item in failures:


            ws.append([

                now,

                "SCAN_FAILED",

                item.get(
                    "Company",
                    ""
                ),

                item.get(
                    "URL",
                    ""
                ),

                item.get(
                    "Error",
                    ""
                )

            ])


        wb.save(
            self.file_path
        )



    def write_run_log(
        self,
        companies_scanned,
        new_jobs,
        updated_jobs,
        failed_companies,
        duration,
        result
    ):


        wb = load_workbook(
            self.file_path
        )


        ws = wb["Run_Log"]


        ws.append([

            now_sgt_str(),

            companies_scanned,

            new_jobs,

            updated_jobs,

            0,

            failed_companies,

            duration,

            result

        ])


        wb.save(
            self.file_path
        )
