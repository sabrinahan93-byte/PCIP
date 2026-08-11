from openpyxl import load_workbook
from pcip.utils.time import now_sgt_str


def generate_html_dashboard(
    excel_path,
    output_path="docs/index.html"
):

    wb = load_workbook(
        excel_path,
        data_only=True
    )

    ws = wb["Jobs"]

    jobs = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if row[0]:

            jobs.append({

                "job_id": row[0],
                "company": row[1],
                "title": row[2],
                "score": row[5],
                "url": row[6],
                "status": row[9] or "",
                "applied_date": row[10] or "",
                "notes": row[11] or ""

            })


    jobs = sorted(
        jobs,
        key=lambda x: x["score"] or 0,
        reverse=True
    )


    high_match = [
        j for j in jobs
        if j["score"]
        and j["score"] >= 75
    ]


    html = f"""

<!DOCTYPE html>

<html>

<head>

<meta charset="UTF-8">

<title>
PCIP Job Dashboard
</title>


<style>

body {{

font-family: Arial;

margin:40px;

background:#f5f5f5;

}}


.card {{

background:white;

padding:20px;

margin-bottom:20px;

border-radius:10px;

}}


table {{

width:100%;

border-collapse:collapse;

}}


td,th {{

padding:10px;

border-bottom:1px solid #ddd;

text-align:left;

}}


.score {{

font-weight:bold;

}}


input, select {{

padding:8px;

margin-right:10px;

margin-bottom:15px;

}}

</style>


</head>


<body>


<h1>
PCIP Daily Job Dashboard
</h1>



<div class="card">

<h2>
GitHub Settings
</h2>


<input
type="password"
id="githubToken"
placeholder="Enter GitHub Token"
>


<button onclick="saveToken()">
Save Token
</button>

<br>


<button onclick="testGitHubConnection()">
Test GitHub Connection
</button>

</div>


<div class="card">

<h2>
Summary
</h2>


<p>
Last Update:
{now_sgt_str()}
</p>


<p>
Total Jobs:
{len(jobs)}
</p>


<p>
High Match Jobs:
{len(high_match)}
</p>


</div>



<div class="card">


<h2>
Match Score >=75
</h2>


<input 
type="text"
id="search"
placeholder="Search Job ID / Company / Title"
onkeyup="filterTable()"
>


<select id="scoreFilter" onchange="filterTable()">

<option value="all">
All Scores
</option>

<option value="75">
>=75
</option>

<option value="85">
>=85
</option>

<option value="90">
>=90
</option>

</select>



<select id="statusFilter" onchange="filterTable()">

<option value="all">
All Status
</option>

<option value="New">
New
</option>

<option value="Watching">
Watching
</option>

<option value="Applied">
Applied
</option>

<option value="Interview">
Interview
</option>

<option value="Rejected">
Rejected
</option>

<option value="Offer">
Offer
</option>

</select>



<table id="jobTable">


<tr>

<th>
Score
</th>

<th>
Company
</th>

<th>
Job Title
</th>

<th>
Pipeline Status
</th>

<th>
Applied Date
</th>

<th>
Notes
</th>

<th>
Link
</th>

</tr>

"""


    for job in high_match:

        html += f"""

<tr>

<td class="score">
{job['score']}
</td>


<td>
{job['company']}
</td>


<td>
{job['title']}
<br>
<small>{job['job_id']}</small>
</td>


<td>

<select
class="status-edit"
data-job-id="{job['job_id']}"
>

<option>
{job['status']}
</option>

<option>
New
</option>

<option>
Watching
</option>

<option>
Applied
</option>

<option>
Interview
</option>

<option>
Rejected
</option>

<option>
Offer
</option>

</select>

</td>


<td>

<input
class="date-edit"
data-job-id="{job['job_id']}"
value="{job['applied_date']}"
>

</td>


<td>

<textarea
class="notes-edit"
data-job-id="{job['job_id']}"
>{job['notes']}</textarea>

</td>


<td>

<a href="{job['url']}">
Open
</a>

</td>


<td>

<button
class="save-btn"
data-job-id="{job['job_id']}"
>
Save
</button>
</td>

</tr>

"""


    html += """

</table>


</div>


<script>

const githubOwner = "sabrinahan93-byte";
const githubRepo = "PCIP";
const githubBranch = "test/dashboard-edit-restore";
const githubFile = "output/update_queue.json";



function saveToken() {

    let token =
    document.getElementById("githubToken").value;


    localStorage.setItem(
        "pcip_token",
        token
    );


    alert(
        "Token saved locally"
    );

}

async function testGitHubConnection(){


    let token =
    localStorage.getItem(
    "pcip_token"
    );


    if (!token){

        alert("No Token");

        return;

    }

    alert(
    "Token exists"
    );


    let url =
    "https://api.github.com/repos/"
    +
    githubOwner
    +
    "/"
    +
    githubRepo
    +
    "/contents/"
    +
    githubFile
    +
    "?ref="
    +
    githubBranch;


    let response =
    await fetch(
        url,
        {
            headers:{
                "Authorization":
                "Bearer "
                +
                token
            }
        }
    );


    if(response.ok){

        let data =
        await response.json();


        alert(
        "GitHub Connection OK | File size: "
        +
        data.size
        );


    }
    else{

        alert(
        "GitHub API Error: "
        +
        response.status
        );

    }

}

    


function filterTable() {

let search =
document.getElementById("search").value.toLowerCase();

let scoreFilter =
document.getElementById("scoreFilter").value;

let statusFilter =
document.getElementById("statusFilter").value;


let table =
document.getElementById("jobTable");


let rows =
table.getElementsByTagName("tr");


for (let i = 1; i < rows.length; i++) {

    let row = rows[i];

    let score =
    row.cells[0].innerText.trim();

    let company =
    row.cells[1].innerText.toLowerCase();

    let title =
    row.cells[2].innerText.toLowerCase();

    let status =
    row.cells[3].querySelector("select").value;


    let matchSearch =
    company.includes(search)
    ||
    title.includes(search);


    let matchScore =
    scoreFilter === ""
    ||
    scoreFilter === "all"
    ||
    parseInt(score) >= parseInt(scoreFilter);


    let matchStatus =
    statusFilter === ""
    ||
    statusFilter === "all"
    ||
    status === statusFilter;


    if(
        matchSearch
        &&
        matchScore
        &&
        matchStatus
    ){

        row.style.display = "";

    }
    else{

        row.style.display = "none";

    }

}

}


// Save button test

// Save button

document.querySelectorAll(".save-btn")
.forEach(button => {

    button.onclick = function(){


        let jobId = this.dataset.jobId;


        let status =
        document.querySelector(
        '.status-edit[data-job-id="' + jobId + '"]'
        ).value;


        let appliedDate =
        document.querySelector(
        '.date-edit[data-job-id="' + jobId + '"]'
        ).value;


        let notes =
        document.querySelector(
        '.notes-edit[data-job-id="' + jobId + '"]'
        ).value;


        let token =
        localStorage.getItem(
        "pcip_token"
        );


        if (!token){

            alert(
            "Please save GitHub Token first"
            );

            return;

        }


        let updateData = {

            job_id: jobId,

            status: status,

            applied_date: appliedDate,

            notes: notes,

            updated_at: new Date().toISOString()

        };


        console.log(
            "Update Request:",
            updateData
        );


        let content =
        btoa(
        JSON.stringify(updateData)
        );


        let url =
        "https://api.github.com/repos/"
        +
        githubOwner
        +
        "/"
        +
        githubRepo
        +
        "/contents/"
        +
        githubFile;



        // Step 1: Get current file SHA

        fetch(
            url,
            {
                headers:{
                    "Authorization":
                    "Bearer "
                    +
                    token,

                    "Accept":
                    "application/vnd.github+json"
                }
            }
        )

        .then(
            response => response.json()
        )

        .then(
            fileData => {


                // Step 2: Update file with SHA


                fetch(
                    url,
                    {

                        method:"PUT",


                        headers:{

                            "Authorization":
                            "Bearer "
                            +
                            token,


                            "Accept":
                            "application/vnd.github+json"

                        },


                        body:JSON.stringify({

                            message:
                            "Update dashboard request",


                            content:
                            content,


                            sha:
                            fileData.sha

                        })

                    }
                )


                .then(
                    response => response.json()
                )


                .then(
                    data => {


                        console.log(
                            "GitHub update result:",
                            data
                        );


                        alert(
                            "Update request uploaded"
                        );


                    }

                );


            }

        );


    };

});

</script>


</body>


</html>

"""


    with open(
        output_path,
        "w",
        encoding="utf-8"
    ) as f:

        f.write(html)
