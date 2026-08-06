from openpyxl import load_workbook
import json
from pathlib import Path


EXCEL_FILE = "output/Job_Dashboard.xlsx"
QUEUE_FILE = "output/update_queue.json"


def update_excel_from_queue():

    queue_path = Path(QUEUE_FILE)

    if not queue_path.exists():
        print("No update queue found")
        return


    with open(queue_path, "r", encoding="utf-8") as f:
        update = json.load(f)



    job_id = update["job_id"]

    status = update.get("status", "")

    applied_date = update.get(
        "applied_date",
        ""
    )

    notes = update.get(
        "notes",
        ""
    )


    wb = load_workbook(
        EXCEL_FILE
    )


    ws = wb["Jobs"]


    updated = False


    for row in ws.iter_rows(
        min_row=2
    ):

        if row[0].value == job_id:


            # Pipeline Status
            row[9].value = status


            # Applied Date
            row[10].value = applied_date


            # Notes
            row[11].value = notes


            updated = True

            print(
                "Updated:",
                job_id
            )

            break



    if updated:

        wb.save(
            EXCEL_FILE
        )

        print(
            "Excel saved"
        )

    else:

        print(
            "Job ID not found:",
            job_id
        )



if __name__ == "__main__":

    update_excel_from_queue()
