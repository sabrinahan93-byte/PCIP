from openpyxl import load_workbook
from pathlib import Path
import re


def normalize(value):

    if value is None:

        return ""

    return (

        str(value)

        .strip()

        .lower()

    )



def extract_job_identifier(url):

    """
    Extract stable job identifier
    Especially for LinkedIn URLs
    """

    if not url:

        return ""


    url = normalize(url)


    # LinkedIn job id
    linkedin_match = re.search(
        r"jobs/view/.*?-(\d+)",
        url
    )


    if linkedin_match:

        return linkedin_match.group(1)



    # Generic URL last number
    number_match = re.search(
        r"(\d+)",
        url
    )


    if number_match:

        return number_match.group(1)



    return url.split("?")[0]



def create_job_key(

    company,

    title,

    url

):


    company = normalize(company)

    title = normalize(title)


    job_id = extract_job_identifier(
        url
    )


    if job_id:

        return (

            company,

            job_id

        )


    return (

        company,

        title

    )



def filter_new_jobs(

    file_path,

    jobs

):


    absolute_path = Path(
        file_path
    ).resolve()


    wb = load_workbook(
        absolute_path
    )


    ws = wb["Jobs"]


    existing_keys = set()



    for row in ws.iter_rows(

        min_row=2,

        values_only=True

    ):


        existing_keys.add(

            create_job_key(

                row[1],

                row[2],

                row[9]

            )

        )



    print(
        "Existing jobs count:",
        len(existing_keys)
    )



    new_jobs = []

    duplicate_count = 0



    for job in jobs:


        key = create_job_key(

            job.get(
                "Company",
                ""
            ),

            job.get(
                "Job Title",
                ""
            ),

            job.get(
                "Job URL",
                ""
            )

        )


        if key in existing_keys:


            duplicate_count += 1


        else:


            print(
                "NEW JOB KEY:",
                key
            )


            new_jobs.append(
                job
            )


            existing_keys.add(
                key
            )



    print(
        "Duplicate jobs:",
        duplicate_count
    )


    print(
        "New jobs after filtering:",
        len(new_jobs)
    )


    return new_jobs
