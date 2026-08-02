from openpyxl import load_workbook
from datetime import datetime
import os


def write_scan_log(
    file_path,
    scan_results,
    duration=None
):

    """
    Release 5.3 Fix

    Purpose:
    Write scanning summary into Sheet 3 Run_Log only.

    IMPORTANT:
    This function MUST NOT write Run_Detail.

    Run_Detail ownership:
    DashboardWorkbook.write_jobs()

    Run_Log ownership:
    write_scan_log()
    """



    if not os.path.exists(file_path):

        raise FileNotFoundError(
            f"Dashboard file not found: {file_path}"
        )



    wb = load_workbook(
        file_path
    )


    if "Run_Log" not in wb.sheetnames:

        raise Exception(
            "Run_Log sheet missing"
        )



    ws = wb["Run_Log"]



    run_time = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )



    companies_scanned = len(
        scan_results
    )



    failed_companies = 0



    for record in scan_results:


        if record.get(
            "Status"
        ) != "Success":

            failed_companies += 1



    result = (

        "Success"

        if failed_companies == 0

        else "Partial"

    )



    #
    # Run_Log only
    #

    ws.append([

        run_time,

        companies_scanned,

        0,          # New Jobs updated elsewhere

        0,          # Updated Jobs updated elsewhere

        0,          # Closed Jobs

        failed_companies,

        duration
        if duration
        else "",

        result

    ])



    wb.save(
        file_path
    )
